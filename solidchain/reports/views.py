from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg
from django.db import models
from datetime import datetime, date, timedelta
import json
from io import BytesIO

from .forms import (
    MonthlyReportForm, 
    MemberStatementForm, 
    OutstandingPaymentsForm, 
    FinesReportForm
)
from core.models import Member, Payment, ContributionMonth
from .models import ReportLog
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

def apply_excel_header_style(ws, row_idx, num_columns):
    """Apply header style to a row in Excel worksheet"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    return ws

def apply_excel_total_style(ws, row_idx, num_columns):
    """Apply total row style in Excel worksheet"""
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = total_font
        cell.fill = total_fill
    
    return ws

def format_currency_column(ws, column_letter, start_row, end_row):
    """Apply currency format to a column range"""
    for row in range(start_row, end_row + 1):
        cell = ws[f"{column_letter}{row}"]
        cell.number_format = '#,##0.00'
    
    return ws



# Helper functions
def calculate_monthly_stats(month):
    """Calculate statistics for a specific month"""
    payments = Payment.objects.filter(month=month).select_related('member__user')
    
    # Get all active members
    all_active_members = Member.objects.filter(is_active=True).select_related('user')
    
    # Get member IDs who have paid for this month
    paid_member_ids = payments.values_list('member_id', flat=True)
    
    # Get pending members (active members who haven't paid)
    pending_members = all_active_members.exclude(id__in=paid_member_ids)
    
    total_members = all_active_members.count()
    paid_members = payments.count()
    on_time_count = payments.filter(status='On Time').count()
    
    # Calculate percentages
    if paid_members > 0:
        on_time_percentage = (on_time_count / paid_members) * 100
        collection_rate = (paid_members / total_members) * 100 if total_members > 0 else 0
    else:
        on_time_percentage = 0
        collection_rate = 0
    
    stats = {
        'month': month,
        'total_members': total_members,
        'paid_members': paid_members,
        'pending_members': pending_members.count(),
        'pending_members_list': pending_members,  # Add this
        'paid_member_ids': list(paid_member_ids),  # Add this for template checks
        'total_collected': payments.aggregate(total=Sum('amount_paid'))['total'] or 0,
        'total_fines': payments.aggregate(total=Sum('fine_amount'))['total'] or 0,
        'on_time_count': on_time_count,
        'late_count': payments.filter(status='Late').count(),
        'average_fine': payments.aggregate(avg=Avg('fine_amount'))['avg'] or 0,
        'collection_rate': collection_rate,
        'on_time_percentage': on_time_percentage,
        'payments': payments,
        'all_active_members': all_active_members,  # Add this
    }
    
    return stats


def get_member_statement(member, start_date=None, end_date=None):
    """Get payment history for a specific member"""
    payments = Payment.objects.filter(member=member).select_related('month')
    
    if start_date:
        payments = payments.filter(month__month__gte=start_date)
    if end_date:
        payments = payments.filter(month__month__lte=end_date)
    
    payments = payments.order_by('-month__month')
    
    total_payments = payments.count()
    on_time_count = payments.filter(status='On Time').count()
    late_count = payments.filter(status='Late').count()
    
    # Calculate percentages
    if total_payments > 0:
        on_time_percentage = (on_time_count / total_payments) * 100
        late_percentage = (late_count / total_payments) * 100
    else:
        on_time_percentage = 0
        late_percentage = 0
    
    summary = {
        'member': member,
        'total_payments': total_payments,
        'total_contributions': payments.aggregate(total=Sum('amount_paid'))['total'] or 0,
        'total_fines': payments.aggregate(total=Sum('fine_amount'))['total'] or 0,
        'on_time_count': on_time_count,
        'late_count': late_count,
        'on_time_percentage': on_time_percentage,
        'late_percentage': late_percentage,  # Add this
        'payments': payments,
        'period_start': start_date,
        'period_end': end_date,
    }
    
    return summary


def get_outstanding_payments(month):
    """Get list of members who haven't paid for a specific month"""
    paid_member_ids = Payment.objects.filter(month=month).values_list('member_id', flat=True)
    outstanding_members = Member.objects.filter(
        is_active=True
    ).exclude(
        id__in=paid_member_ids
    ).select_related('user').order_by('user__first_name')
    
    paid_count = Payment.objects.filter(month=month).count()
    total_members = Member.objects.filter(is_active=True).count()
    outstanding_count = outstanding_members.count()
    
    # Calculate collection rate
    if total_members > 0:
        collection_rate = (paid_count / total_members) * 100
    else:
        collection_rate = 0
    
    return {
        'month': month,
        'outstanding_members': outstanding_members,
        'count': outstanding_count,
        'paid_count': paid_count,  # Add this
        'total_members': total_members,  # Add this
        'collection_rate': collection_rate,  # Add this
        'total_amount': outstanding_count * 2500,  # Monthly contribution
    }


def get_fines_summary(start_date=None, end_date=None, group_by_month=True):
    """Get fines summary with optional grouping"""
    payments = Payment.objects.filter(fine_amount__gt=0).select_related('member__user', 'month')
    
    if start_date:
        payments = payments.filter(month__month__gte=start_date)
    if end_date:
        payments = payments.filter(month__month__lte=end_date)
    
    total_fines = payments.aggregate(total=Sum('fine_amount'))['total'] or 0
    total_fine_payments = payments.count()
    
    if group_by_month:
        # Group fines by month
        monthly_fines = payments.values(
            'month__month', 
            'month__month__year', 
            'month__month__month'
        ).annotate(
            total_fines=Sum('fine_amount'),
            count=Count('id'),
            avg_fine=Avg('fine_amount')
        ).order_by('-month__month')
        
        return {
            'total_fines': total_fines,
            'total_fine_payments': total_fine_payments,
            'monthly_fines': monthly_fines,
            'payments': payments.order_by('-month__month'),
            'average_fine': total_fines / total_fine_payments if total_fine_payments > 0 else 0,
            'start_date': start_date,
            'end_date': end_date,
        }
    else:
        return {
            'total_fines': total_fines,
            'total_fine_payments': total_fine_payments,
            'payments': payments.order_by('-month__month'),
            'average_fine': total_fines / total_fine_payments if total_fine_payments > 0 else 0,
            'start_date': start_date,
            'end_date': end_date,
        }
# Main views
@staff_member_required
def reports_dashboard(request):
    """Main reports dashboard"""
    # Get some quick stats for the dashboard
    current_month = timezone.now().date().replace(day=1)
    
    try:
        current_contribution_month = ContributionMonth.objects.get(month=current_month)
        current_stats = calculate_monthly_stats(current_contribution_month)
    except ContributionMonth.DoesNotExist:
        current_stats = None
    
    # Recent report logs
    recent_reports = ReportLog.objects.filter(
        generated_by=request.user
    ).order_by('-generated_at')[:5]
    
    context = {
        'current_stats': current_stats,
        'recent_reports': recent_reports,
        'total_members': Member.objects.filter(is_active=True).count(),
        'total_payments': Payment.objects.count(),
        'total_fines_collected': Payment.objects.aggregate(total=Sum('fine_amount'))['total'] or 0,
    }
    
    return render(request, 'reports/dashboard.html', context)


@staff_member_required
def monthly_collection_report(request):
    """Generate monthly collection report"""
    if request.method == 'POST':
        form = MonthlyReportForm(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']
            include_pending = form.cleaned_data['include_pending']
            show_fine_details = form.cleaned_data['show_fine_details']
            format_type = form.cleaned_data['format']
            
            # Get statistics for the month
            stats = calculate_monthly_stats(month)
            print(stats)
            # Log the report generation
            ReportLog.objects.create(
                report_type='monthly',
                format=format_type,
                generated_by=request.user,
                parameters={
                    'month': month.id,
                    'month_name': str(month),
                    'include_pending': include_pending,
                    'show_fine_details': show_fine_details,
                }
            )
            
            # Prepare context
            context = {
                'stats': stats,
                'month': month,
                'include_pending': include_pending,
                'show_fine_details': show_fine_details,
                'generated_at': timezone.now(),
                'generated_by': request.user,
                'report_title': f'Monthly Collection Report - {month.month.strftime("%B %Y")}',
            }
            
            # Handle different formats
            if format_type == 'pdf':
                return generate_pdf_report('reports/monthly_pdf.html', context, f'monthly_report_{month.month.strftime("%Y_%m")}')
            elif format_type == 'excel':
                return generate_excel_monthly_report(stats, month)
            else:
                # HTML format - render template
                return render(request, 'reports/monthly_report.html', context)
    else:
        form = MonthlyReportForm()
    
    # Get recent months for quick links
    recent_months = ContributionMonth.objects.all().order_by('-month')[:6]
    
    context = {
        'form': form,
        'recent_months': recent_months,
        'report_type': 'Monthly Collection',
    }
    return render(request, 'reports/report_form.html', context)


@staff_member_required
def member_statement(request):
    """Generate member statement report"""
    if request.method == 'POST':
        form = MemberStatementForm(request.POST)
        if form.is_valid():
            member = form.cleaned_data['member']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            include_summary = form.cleaned_data['include_summary']
            format_type = form.cleaned_data['format']
            
            # Get member statement
            statement = get_member_statement(member, start_date, end_date)
            
            # Log the report generation
            ReportLog.objects.create(
                report_type='member',
                format=format_type,
                generated_by=request.user,
                parameters={
                    'member_id': member.id,
                    'member_name': str(member),
                    'start_date': str(start_date) if start_date else None,
                    'end_date': str(end_date) if end_date else None,
                }
            )
            
            # Prepare context
            context = {
                'statement': statement,
                'member': member,
                'start_date': start_date,
                'end_date': end_date,
                'include_summary': include_summary,
                'generated_at': timezone.now(),
                'generated_by': request.user,
                'report_title': f'Member Statement - {member.user.get_full_name()}',
            }
            
            # Handle different formats
            if format_type == 'pdf':
                return generate_pdf_report('reports/member_pdf.html', context, f'member_statement_{member.id}_{datetime.now().strftime("%Y%m%d")}')
            elif format_type == 'excel':
                return generate_excel_member_report(statement, member)
            else:
                # HTML format - render template
                return render(request, 'reports/member_statement.html', context)
    else:
        form = MemberStatementForm()
    
    # Get active members for quick links
    active_members = Member.objects.filter(is_active=True).select_related('user').order_by('user__first_name')[:10]
    
    context = {
        'form': form,
        'active_members': active_members,
        'report_type': 'Member Statement',
    }
    return render(request, 'reports/report_form.html', context)


@staff_member_required
def outstanding_payments_report(request):
    """Generate outstanding payments report"""
    if request.method == 'POST':
        form = OutstandingPaymentsForm(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']
            send_reminders = form.cleaned_data['send_reminders']
            format_type = form.cleaned_data['format']
            
            # Get outstanding payments
            outstanding_data = get_outstanding_payments(month)
            
            # Log the report generation
            ReportLog.objects.create(
                report_type='outstanding',
                format=format_type,
                generated_by=request.user,
                parameters={
                    'month': month.id,
                    'month_name': str(month),
                    'send_reminders': send_reminders,
                }
            )
            
            # Send reminders if requested
            if send_reminders:
                # This would integrate with your notification system
                messages.info(request, f'Reminder notes prepared for {outstanding_data["count"]} members.')
            
            # Prepare context
            context = {
                'outstanding_data': outstanding_data,
                'month': month,
                'send_reminders': send_reminders,
                'generated_at': timezone.now(),
                'generated_by': request.user,
                'report_title': f'Outstanding Payments - {month.month.strftime("%B %Y")}',
            }
            
            # Handle different formats
            if format_type == 'pdf':
                return generate_pdf_report('reports/outstanding_pdf.html', context, f'outstanding_payments_{month.month.strftime("%Y_%m")}')
            elif format_type == 'excel':
                return generate_excel_outstanding_report(outstanding_data, month)
            else:
                # HTML format - render template
                return render(request, 'reports/outstanding_payments.html', context)
    else:
        form = OutstandingPaymentsForm()
    
    # Get current month for default
    current_month = timezone.now().date().replace(day=1)
    try:
        default_month = ContributionMonth.objects.get(month=current_month)
    except ContributionMonth.DoesNotExist:
        default_month = None
    
    context = {
        'form': form,
        'default_month': default_month,
        'report_type': 'Outstanding Payments',
    }
    return render(request, 'reports/report_form.html', context)


@staff_member_required
def fines_summary_report(request):
    """Generate fines summary report"""
    if request.method == 'POST':
        form = FinesReportForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            group_by_month = form.cleaned_data['group_by_month']
            format_type = form.cleaned_data['format']
            
            # Get fines summary
            fines_data = get_fines_summary(start_date, end_date, group_by_month)
            
            # Log the report generation
            ReportLog.objects.create(
                report_type='fines',
                format=format_type,
                generated_by=request.user,
                parameters={
                    'start_date': str(start_date) if start_date else None,
                    'end_date': str(end_date) if end_date else None,
                    'group_by_month': group_by_month,
                }
            )
            
            # Prepare context
            context = {
                'fines_data': fines_data,
                'start_date': start_date,
                'end_date': end_date,
                'group_by_month': group_by_month,
                'generated_at': timezone.now(),
                'generated_by': request.user,
                'report_title': 'Fines Summary Report',
            }
            
            if start_date and end_date:
                context['report_title'] = f'Fines Summary ({start_date.strftime("%d/%m/%Y")} to {end_date.strftime("%d/%m/%Y")})'
            
            # Handle different formats
            if format_type == 'pdf':
                return generate_pdf_report('reports/fines_pdf.html', context, f'fines_summary_{datetime.now().strftime("%Y%m%d")}')
            elif format_type == 'excel':
                return generate_excel_fines_report(fines_data)
            else:
                # HTML format - render template
                return render(request, 'reports/fines_summary.html', context)
    else:
        # Set default dates (last 6 months)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=180)
        
        form = FinesReportForm(initial={
            'start_date': start_date,
            'end_date': end_date,
            'group_by_month': True,
        })
    
    context = {
        'form': form,
        'report_type': 'Fines Summary',
    }
    return render(request, 'reports/report_form.html', context)


@staff_member_required
def report_history(request):
    """View report generation history"""
    reports = ReportLog.objects.filter(generated_by=request.user).order_by('-generated_at')
    
    # Stats
    total_reports = reports.count()
    reports_today = reports.filter(generated_at__date=timezone.now().date()).count()
    
    context = {
        'reports': reports,
        'total_reports': total_reports,
        'reports_today': reports_today,
    }
    
    return render(request, 'reports/history.html', context)


# Export functions (simplified versions - you'll need to implement full versions)

def generate_pdf_report(template_name, context, filename):
    """Generate PDF report - placeholder function"""
    # You'll need to install reportlab or weasyprint for this
    messages.warning(request, 'PDF export is not yet implemented. Please use HTML or Excel format.')
    return redirect('reports:dashboard')



@staff_member_required
def generate_excel_monthly_report(stats, month):
    """Generate Excel monthly report - Works with QuerySet or list of dictionaries"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly Report {month.month.strftime('%Y-%m')}"
    
    # Set column widths
    column_widths = [30, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Title
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = f"SOLIDCHAIN - MONTHLY COLLECTION REPORT - {month.month.strftime('%B %Y')}"
    title.font = Font(bold=True, size=14, color="4361ee")
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Report Info
    ws.append([])
    ws.append(["For Month:", month.month.strftime("%B %Y")])
    ws.append(["Due Date:", month.due_date.strftime("%Y-%m-%d")])
    ws.append(["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Generated By:", str(stats.get('generated_by', 'System'))])
    ws.append([])
    
    # Summary Statistics
    ws.append(["COLLECTION SUMMARY"])
    ws.append([])
    
    summary_data = [
        ["Total Active Members", stats.get('total_members', 0)],
        ["Members Paid", stats.get('paid_members', 0)],
        ["Members Pending", stats.get('pending_members', 0)],
        ["Collection Rate", f"{stats.get('collection_rate', 0):.1f}%"],
        ["Total Amount Collected", f"KSh {stats.get('total_collected', 0):,.2f}"],
        ["Total Fines Collected", f"KSh {stats.get('total_fines', 0):,.2f}"],
        ["On Time Payments", stats.get('on_time_count', 0)],
        ["Late Payments", stats.get('late_count', 0)],
        ["On Time Percentage", f"{stats.get('on_time_percentage', 0):.1f}%"],
    ]
    
    for row in summary_data:
        ws.append(row)
    
    ws.append([])
    
    # Payment Details Header
    headers = ["Member Name", "Phone", "Amount Paid", "Fine Amount", "Status", "Paid Date"]
    ws.append(["PAYMENT DETAILS"])
    ws.append([])
    ws.append(headers)
    
    # Header styles
    ws = apply_excel_header_style(ws, ws.max_row, len(headers))
    
    # Payment Details Data
    payments = stats.get('payments', [])
    
    # Ensure we're working with a list
    if hasattr(payments, 'query'):  # It's a QuerySet
        # Evaluate the QuerySet to get model instances
        payments = list(payments)
    
    for payment in payments:
        try:
            # Handle both model instances and dictionaries
            if isinstance(payment, dict):
                # Extract data from dictionary
                member_name = payment.get('member__user__first_name', '') + " " + payment.get('member__user__last_name', '')
                member_name = member_name.strip() or payment.get('member__user__username', 'Unknown')
                
                # Try different possible keys for phone
                phone = (payment.get('member__phone') or 
                        payment.get('phone') or 
                        "N/A")
                
                amount = float(payment.get('amount_paid', 0))
                fine = float(payment.get('fine_amount', 0))
                status = payment.get('status', 'Unknown')
                
                paid_date_str = payment.get('paid_date')
                if paid_date_str and hasattr(paid_date_str, 'strftime'):
                    paid_date = paid_date_str.strftime("%Y-%m-%d")
                else:
                    paid_date = str(paid_date_str) if paid_date_str else ""
                    
            else:
                # Assume it's a model instance
                # Get member name
                if hasattr(payment, 'member'):
                    member = payment.member
                    if hasattr(member, 'user'):
                        user = member.user
                        if hasattr(user, 'get_full_name'):
                            full_name = user.get_full_name()
                            member_name = full_name if full_name else user.username
                        else:
                            member_name = user.username
                    else:
                        member_name = str(member)
                else:
                    member_name = "Unknown"
                
                # Get phone
                phone = getattr(payment.member, 'phone', 'N/A') if hasattr(payment, 'member') else 'N/A'
                
                # Get payment details
                amount = float(getattr(payment, 'amount_paid', 0))
                fine = float(getattr(payment, 'fine_amount', 0))
                status = getattr(payment, 'status', 'Unknown')
                
                # Get paid date
                paid_date_obj = getattr(payment, 'paid_date', None)
                if paid_date_obj and hasattr(paid_date_obj, 'strftime'):
                    paid_date = paid_date_obj.strftime("%Y-%m-%d")
                else:
                    paid_date = ""
            
            ws.append([
                member_name,
                phone,
                amount,
                fine,
                status,
                paid_date
            ])
            
        except Exception as e:
            print(f"Error processing payment: {e}")
            # Add error row for debugging
            ws.append([f"Error: {str(e)[:50]}", "", "", "", "", ""])
    
    # Format currency columns (Amount and Fine columns)
    num_rows = len(payments) + 1  # +1 for header row
    if num_rows > 1:
        # Apply currency format to Amount column (column C)
        for row in range(ws.max_row - len(payments) + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)  # Column C
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=row, column=4)  # Column D (Fine)
            cell.number_format = '#,##0.00'
    
    # Add totals row
    ws.append([])
    total_row = ["TOTAL", "", 
                stats.get('total_collected', 0),
                stats.get('total_fines', 0),
                "", ""]
    ws.append(total_row)
    
    # Style totals row
    ws = apply_excel_total_style(ws, ws.max_row, len(total_row))
    
    # Format totals row as currency
    ws.cell(row=ws.max_row, column=3).number_format = '#,##0.00'
    ws.cell(row=ws.max_row, column=4).number_format = '#,##0.00'
    
    # Pending Members Section (if needed)
    if stats.get('pending_members', 0) > 0:
        ws.append([])
        ws.append(["PENDING MEMBERS"])
        ws.append([])
        pending_headers = ["Member Name", "Phone", "Email", "Member Since"]
        ws.append(pending_headers)
        
        # Header style for pending section
        ws = apply_excel_header_style(ws, ws.max_row, len(pending_headers))
        
        # Get pending members - handle both QuerySet and list
        pending_members = stats.get('pending_members_list', [])
        
        if hasattr(pending_members, 'query'):  # It's a QuerySet
            pending_members = list(pending_members)
        
        for member in pending_members:
            try:
                if isinstance(member, dict):
                    member_name = member.get('user__first_name', '') + " " + member.get('user__last_name', '')
                    member_name = member_name.strip() or member.get('user__username', 'Unknown')
                    phone = member.get('phone', 'N/A')
                    email = member.get('user__email', 'N/A')
                    member_since = member.get('joined_date', '')
                else:
                    # Model instance
                    if hasattr(member, 'user'):
                        user = member.user
                        member_name = user.get_full_name() or user.username
                        email = user.email or "N/A"
                    else:
                        member_name = str(member)
                        email = "N/A"
                    
                    phone = getattr(member, 'phone', 'N/A')
                    member_since = getattr(member, 'joined_date', '').strftime("%Y-%m-%d") if hasattr(getattr(member, 'joined_date', ''), 'strftime') else str(getattr(member, 'joined_date', ''))
                
                ws.append([member_name, phone, email, member_since])
            except Exception as e:
                print(f"Error processing pending member: {e}")
                ws.append([f"Error: {str(e)[:30]}", "", "", ""])
        
        ws.append([])
        pending_total = ["TOTAL PENDING", f"{stats.get('pending_members', 0)} members", "", ""]
        ws.append(pending_total)
        ws = apply_excel_total_style(ws, ws.max_row, len(pending_total))
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"monthly_report_{month.month.strftime('%Y_%m')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response
@staff_member_required
def generate_excel_member_report(statement, member):
    """Generate Excel member statement report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Member Statement {member.id}"
    
    # Column widths
    column_widths = [25, 15, 15, 15, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Title
    ws.merge_cells('A1:G1')
    title = ws['A1']
    title.value = f"SOLIDCHAIN - MEMBER STATEMENT - {member.user.get_full_name() or member.user.username}"
    title.font = Font(bold=True, size=14, color="4361ee")
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Member Info
    ws.append([])
    ws.append(["Member ID:", member.id])
    ws.append(["Full Name:", member.user.get_full_name() or member.user.username])
    ws.append(["Phone:", member.phone or "N/A"])
    ws.append(["Email:", member.user.email or "N/A"])
    ws.append(["Member Since:", member.joined_date.strftime("%Y-%m-%d")])
    ws.append(["Status:", "Active" if member.is_active else "Inactive"])
    
    ws.append([])
    
    # Statement Summary
    ws.append(["STATEMENT SUMMARY"])
    ws.append([])
    
    summary_data = [
        ["Total Payments", statement['total_payments']],
        ["Total Contributions", f"KSh {statement['total_contributions']:,.2f}"],
        ["Total Fines Paid", f"KSh {statement['total_fines']:,.2f}"],
        ["Total Amount Paid", f"KSh {statement['total_contributions'] + statement['total_fines']:,.2f}"],
        ["On Time Payments", statement['on_time_count']],
        ["Late Payments", statement['late_count']],
        ["On Time Rate", f"{statement.get('on_time_percentage', 0):.1f}%"],
    ]
    
    for row in summary_data:
        ws.append(row)
    
    ws.append([])
    
    # Payment History Header
    headers = ["Month", "Paid Date", "Due Date", "Amount", "Fine", "Status", "Days Late"]
    ws.append(["PAYMENT HISTORY"])
    ws.append([])
    ws.append(headers)
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Payment History Data
    for payment in statement['payments']:
        month = payment.month.month.strftime("%B %Y")
        paid_date = payment.paid_date.strftime("%Y-%m-%d")
        due_date = payment.month.due_date.strftime("%Y-%m-%d")
        amount = f"KSh {payment.amount_paid:,.2f}"
        fine = f"KSh {payment.fine_amount:,.2f}" if payment.fine_amount > 0 else "None"
        status = payment.status
        days_late = payment.days_late if hasattr(payment, 'days_late') else 0
        
        ws.append([month, paid_date, due_date, amount, fine, status, days_late])
    
    # Format currency columns
    for row in ws.iter_rows(min_row=ws.max_row - statement['payments'].count() + 1, max_row=ws.max_row):
        row[3].number_format = '#,##0.00'  # Amount
        row[4].number_format = '#,##0.00'  # Fine
    
    # Totals
    ws.append([])
    total_row = ["TOTAL", "", "", f"KSh {statement['total_contributions']:,.2f}", 
                f"KSh {statement['total_fines']:,.2f}", "", ""]
    ws.append(total_row)
    
    # Style totals
    totals_font = Font(bold=True, color="FFFFFF")
    totals_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    for col in range(1, len(total_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = totals_font
        cell.fill = totals_fill
    
    # Report footer
    ws.append([])
    ws.append(["Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if statement['period_start']:
        ws.append(["Period Start:", statement['period_start'].strftime("%Y-%m-%d")])
    if statement['period_end']:
        ws.append(["Period End:", statement['period_end'].strftime("%Y-%m-%d")])
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"member_statement_{member.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response


@staff_member_required
def generate_excel_outstanding_report(outstanding_data, month):
    """Generate Excel outstanding payments report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Outstanding Payments {month.month.strftime('%Y-%m')}"
    
    # Column widths
    column_widths = [30, 20, 20, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Title
    ws.merge_cells('A1:E1')
    title = ws['A1']
    title.value = f"SOLIDCHAIN - OUTSTANDING PAYMENTS - {month.month.strftime('%B %Y')}"
    title.font = Font(bold=True, size=14, color="dc3545")
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Report Info
    ws.append([])
    ws.append(["For Month:", month.month.strftime("%B %Y")])
    ws.append(["Due Date:", month.due_date.strftime("%Y-%m-%d")])
    ws.append(["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    
    # Summary Statistics
    ws.append(["SUMMARY STATISTICS"])
    ws.append([])
    
    summary_data = [
        ["Total Members", outstanding_data['total_members']],
        ["Paid Members", outstanding_data['paid_count']],
        ["Outstanding Members", outstanding_data['count']],
        ["Collection Rate", f"{outstanding_data.get('collection_rate', 0):.1f}%"],
        ["Total Amount Outstanding", f"KSh {outstanding_data['total_amount']:,.2f}"],
        ["Monthly Contribution", "KSh 2,500.00"],
    ]
    
    for row in summary_data:
        ws.append(row)
    
    ws.append([])
    
    # Outstanding Members Header
    headers = ["Member Name", "Phone", "Email", "Member Since", "Amount Due"]
    ws.append(["OUTSTANDING MEMBERS"])
    ws.append([])
    ws.append(headers)
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Outstanding Members Data
    for member in outstanding_data['outstanding_members']:
        member_name = member.user.get_full_name() or member.user.username
        phone = member.phone or "N/A"
        email = member.user.email or "N/A"
        member_since = member.joined_date.strftime("%Y-%m-%d")
        amount_due = "KSh 2,500.00"
        
        ws.append([member_name, phone, email, member_since, amount_due])
    
    # Totals
    ws.append([])
    total_row = ["TOTAL OUTSTANDING", "", "", "", f"KSh {outstanding_data['total_amount']:,.2f}"]
    ws.append(total_row)
    
    # Style totals
    totals_font = Font(bold=True, color="FFFFFF")
    totals_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    for col in range(1, len(total_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = totals_font
        cell.fill = totals_fill
    
    # Reminder Notes
    ws.append([])
    ws.append(["REMINDER NOTES"])
    ws.append([])
    reminder_notes = [
        "1. Payment was due on " + month.due_date.strftime("%B %d, %Y"),
        "2. Late payments incur fines as follows:",
        "   - 6th to 10th: KSh 100 per day",
        "   - 11th onwards: KSh 25 per day",
        "3. Fines stop accumulating on the 5th of the following month",
        "4. Please make payment to the group treasurer",
        "5. Contact treasurer for payment details",
    ]
    
    for note in reminder_notes:
        ws.append([note])
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"outstanding_payments_{month.month.strftime('%Y_%m')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

@staff_member_required
def generate_excel_fines_report(fines_data):
    """Generate Excel fines summary report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fines Summary"
    
    # Title
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = "SOLIDCHAIN - FINES SUMMARY REPORT"
    title.font = Font(bold=True, size=14, color="dc3545")
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Report Info
    ws.append([])
    ws.append(["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if fines_data.get('start_date'):
        ws.append(["Start Date:", fines_data['start_date'].strftime("%Y-%m-%d")])
    if fines_data.get('end_date'):
        ws.append(["End Date:", fines_data['end_date'].strftime("%Y-%m-%d")])
    ws.append([])
    
    # Summary Statistics
    ws.append(["SUMMARY STATISTICS"])
    ws.append([])
    
    avg_fine = fines_data['total_fines'] / fines_data['total_fine_payments'] if fines_data['total_fine_payments'] > 0 else 0
    
    summary_data = [
        ["Total Fines Collected", f"KSh {fines_data['total_fines']:,.2f}"],
        ["Total Fine Payments", fines_data['total_fine_payments']],
        ["Average Fine", f"KSh {avg_fine:,.2f}"],
        ["Report Period", f"{fines_data.get('start_date', 'All time')} to {fines_data.get('end_date', 'Present')}"],
    ]
    
    for row in summary_data:
        ws.append(row)
    
    ws.append([])
    
    # Check if we have monthly data or detailed payments
    if fines_data.get('monthly_fines'):
        # Monthly Breakdown
        ws.append(["MONTHLY BREAKDOWN"])
        ws.append([])
        
        headers = ["Month", "Total Fines", "Number of Fines", "Average Fine"]
        ws.append(headers)
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Monthly data
        for month_data in fines_data['monthly_fines']:
            month_str = f"{month_data['month__month__year']}-{month_data['month__month__month']:02d}"
            total_fines = f"KSh {month_data['total_fines']:,.2f}"
            count = month_data['count']
            avg = f"KSh {month_data.get('avg_fine', 0):,.2f}"
            
            ws.append([month_str, total_fines, count, avg])
        
        # Format columns
        for row in ws.iter_rows(min_row=ws.max_row - len(fines_data['monthly_fines']) + 1, max_row=ws.max_row):
            row[1].number_format = '#,##0.00'  # Total fines
            row[3].number_format = '#,##0.00'  # Average fine
    
    elif fines_data.get('payments'):
        # Detailed Payments
        ws.append(["DETAILED FINE PAYMENTS"])
        ws.append([])
        
        headers = ["Member", "Month", "Paid Date", "Days Late", "Fine Amount", "Status"]
        ws.append(headers)
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Payment data
        for payment in fines_data['payments']:
            member_name = payment.member.user.get_full_name() or payment.member.user.username
            month = payment.month.month.strftime("%B %Y")
            paid_date = payment.paid_date.strftime("%Y-%m-%d")
            days_late = payment.days_late if hasattr(payment, 'days_late') else 0
            fine_amount = f"KSh {payment.fine_amount:,.2f}"
            status = payment.status
            
            ws.append([member_name, month, paid_date, days_late, fine_amount, status])
        
        # Format currency column
        for row in ws.iter_rows(min_row=ws.max_row - fines_data['payments'].count() + 1, max_row=ws.max_row):
            row[4].number_format = '#,##0.00'  # Fine amount
    
    # Fine Rules
    ws.append([])
    ws.append(["FINE RULES"])
    ws.append([])
    rules = [
        "1. Due Date: 5th of every month",
        "2. Fine Structure:",
        "   - Days 1-5 late (6th-10th): KSh 100 per day",
        "   - Days 6+ late (11th+): KSh 25 per day",
        "   - Maximum fine period: Until 5th of following month",
        "3. Example: Payment on 15th (10 days late):",
        "   - First 5 days: 5 × KSh 100 = KSh 500",
        "   - Next 5 days: 5 × KSh 25 = KSh 125",
        "   - Total Fine: KSh 625",
    ]
    
    for rule in rules:
        ws.append([rule])
    
    # Set column widths
    column_widths = [25, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"fines_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response

@staff_member_required
def quick_monthly_report(request, month_id):
    """Quick access to monthly report for a specific month"""
    month = get_object_or_404(ContributionMonth, id=month_id)
    stats = calculate_monthly_stats(month)
    
    context = {
        'stats': stats,
        'month': month,
        'include_pending': True,
        'show_fine_details': True,
        'generated_at': timezone.now(),
        'generated_by': request.user,
        'report_title': f'Monthly Collection Report - {month.month.strftime("%B %Y")}',
        'is_quick_view': True,
    }
    
    return render(request, 'reports/monthly_report.html', context)


@staff_member_required 
def quick_member_statement(request, member_id):
    """Quick access to member statement"""
    member = get_object_or_404(Member, id=member_id)
    statement = get_member_statement(member)
    
    context = {
        'statement': statement,
        'member': member,
        'include_summary': True,
        'generated_at': timezone.now(),
        'generated_by': request.user,
        'report_title': f'Member Statement - {member.user.get_full_name()}',
        'is_quick_view': True,
    }
    
    return render(request, 'reports/member_statement.html', context)